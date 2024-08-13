import json
import os
import re
import logging
import time
import concurrent.futures
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime


class ExcelOperations:
	errors = []
	verified_emails = set()

	def __init__(self):
		self.params = self.load_params()

	def load_params(self):
		"""
		This function loads the parameters from the JSON file.
		"""
		try:
			with open("./config/setting_excel.json", 'r') as f:
				return json.load(f)
		except FileNotFoundError:
			logging.error('json file not found')
			return None

	def process_excel_files(self, folder_path, progress_bar=None):
		"""
		This function processes all Excel files in the given folder path and extracts information from them.
		folder_path: path to the folder containing Excel files to process
		progress_bar: tkinter progress bar to update progress
		"""
		information = []
		excel_files = self.get_excel_files(folder_path)
		total_files = len(excel_files)

		if progress_bar:
			progress_bar['value'] = 5
			progress_bar.update()

		for index, (file_path, regex) in enumerate(excel_files):
			try:
				info_from_excel = self.extract_information_from_excel(file_path, regex)

				if info_from_excel is None:
					continue

				if any(info_from_excel['last_name'] == item['last_name'] and
				       info_from_excel['first_name'] == item['first_name'] for item in information):
					continue
				interviews = []
				for date, manager in zip(info_from_excel['dates'], info_from_excel['managers']):
					interviews.append({'date': date, 'manager': manager})
				info_from_excel['interviews'] = interviews
				information.append(info_from_excel)
			except Exception as e:
				logging.error(f"Error while processing file {file_path}: {e}")

			if progress_bar:
				progress_value = ((index + 1) / total_files) * 100
				progress_bar['value'] = progress_value
				progress_bar.update()

		if progress_bar:
			progress_bar['value'] = 100
			progress_bar.update()

		if not information:
			return None

		# Show all accumulated errors at the end of processing
		if self.errors:
			error_message = ", ".join(self.errors) + " ne respectent pas le format d'email mais quand même on l'a " \
			                                         "ajouté dans le nouveau fichier excel. Veuillez les corriger."
			messagebox.showerror("Errors", error_message)

		return information

	def get_excel_files(self, folder_path):
		"""
		This function returns a list of Excel files in the given folder path.
		folder_path: path to the folder containing Excel files
		"""
		excel_files = []
		with concurrent.futures.ThreadPoolExecutor() as executor:
			futures = []
			for root, dirs, files in os.walk(folder_path):
				for file in files:
					if file.endswith(".xlsx") or file.endswith(".xls"):
						file_path = os.path.join(root, file)
						futures.append(executor.submit(self.check_file, file_path, root))

			for future in concurrent.futures.as_completed(futures):
				result = future.result()
				if result:
					excel_files.append(result)

		return excel_files

	def check_file(self, file_path, root):
		"""
		This function checks if the given file is an Excel file, if the format is correct and if it is open,
		file_path: path to the file to check
		root: root directory of the file
		"""
		if not os.path.exists(file_path):
			return None
		try:
			is_open = self.is_file_open(file_path)
			check_format = self.verify_format_file_excel(file_path)
			if is_open:
				if check_format:
					return file_path, self.extract_path_profile(root)
				else:
					return None
			else:
				if check_format:
					return file_path, self.extract_path_profile(root)
		except PermissionError:
			time.sleep(1)
			logging.error("Error ExcelOprt: check_file > no permision")
		except Exception as e:
			logging.error(f"Error ExcelOprt: check_file: {file_path}: {e}")
			return None
		return None

	def extract_contact_information(self, sheet, params):
		"""
		This function extracts contact information from the given Excel sheet.
		sheet: Excel sheet to extract information from
		params: dictionary containing cell references for contact information
		"""
		tel_num_cell = params.get("tel_num")
		tel_num = sheet[tel_num_cell].value
		if tel_num:
			tel_num = re.sub(r'\D', '', tel_num)
		tel_num_sec_cell = params.get("tel_num_sec")
		tel_num_sec = sheet[tel_num_sec_cell].value
		email_cell = params.get("email")
		email_raw = sheet[email_cell].value.strip().replace("Mail: ", "") if sheet[email_cell].value else None
		email = self.clean_email(email_raw)
		data_contact = [tel_num, email]

		if not self.verify_value(data_contact):
			if tel_num_sec:
				tel_num_sec = re.sub(r'\D', '', tel_num_sec)
				data_contact = [tel_num_sec, email]
			if self.verify_value(data_contact):
				tel_num = data_contact[0]

		if tel_num:
			tel_num = ' '.join([tel_num[i:i + 2] for i in range(0, len(tel_num), 2)])
		return tel_num, email

	def cell_reference_to_row(self, cell_ref):
		# Convert cell reference (e.g., 'H7') to row index (e.g., 7)
		return int(re.findall(r'\d+', cell_ref)[0])

	def extract_interview_information(self, sheet, params):
		"""
		This function extracts interview information from the given Excel sheet.
		sheet: Excel sheet to extract information from
		params: dictionary containing cell references for interview information
		"""
		dates = []
		managers = []

		# Convert cell references to row indices for interview dates
		start_row = self.cell_reference_to_row(params["interview_dates_start"])
		end_row = self.cell_reference_to_row(params["interview_dates_end"])
		if not isinstance(start_row, int) or not isinstance(end_row, int):
			raise ValueError(
				"Les valeurs de 'interview_dates_start' et 'interview_dates_end' doivent être des entiers.")

		for row in range(start_row, end_row + 1):
			date_cell = sheet.cell(row=row, column=8).value
			date_value = self.verify_format_date(date_cell)
			if date_value:
				date_value = self.format_date(date_value)
			dates.append(date_value)

		# Convert cell references to row indices for interview managers
		start_row = self.cell_reference_to_row(params["interview_managers_start"])
		end_row = self.cell_reference_to_row(params["interview_managers_end"])
		if not isinstance(start_row, int) or not isinstance(end_row, int):
			raise ValueError(
				"Les valeurs de 'interview_managers_start' et 'interview_managers_end' doivent être des entiers.")

		for row in range(start_row, end_row + 1):
			manager_cell = sheet.cell(row=row, column=7).value
			managers.append(manager_cell if manager_cell else "")

		last_interview = self.find_last_interview(dates)
		return dates, managers, last_interview

	def extract_information_from_excel(self, file_path, regex):
		"""
		This function extracts information from the given Excel file.
		file_path: path to the Excel file to extract information from
		regex: regex pattern to extract profile information from the file path
		"""
		wb = load_workbook(file_path)
		sheet = wb.active

		tel_num, email = self.extract_contact_information(sheet, self.params)
		dates, managers, last_interview = self.extract_interview_information(sheet, self.params)

		profile = regex[0] if regex else ''
		return {
			'profile': profile,
			'direction': file_path,
			'last_name': sheet[self.params["last_name"]].value,
			'first_name': sheet[self.params["first_name"]].value,
			'tel_num': tel_num,
			'email': email,
			'status': sheet[self.params["status1"]].value or sheet[self.params["status2"]].value,
			'dates': dates,
			'managers': managers,
			'last_interview': last_interview
		}

	def create_headers(self):
		"""
		This function creates headers for the new Excel file.
		"""
		headers = ["REPARTOIRE", "PROFIL", "NOM", "PRENOM", "TEL", "EMAIL", "DISPONIBILITE"]
		for i in range(3):
			headers.append(f"DATE {i + 1}")
			headers.append(f"ENTRETIEN {i + 1}")
		headers.append("DERNIER ENTRETIEN")
		return headers

	def create_new_excel_file(self, information, output_path):
		"""
		This function creates a new Excel file with the given information.
		information: list of dictionaries containing information to write to the new Excel file
		output_path: path to the new Excel file
		"""
		if not information:
			raise ValueError("No information to write, file would be empty except headers")

		wb = Workbook()
		sheet = wb.active
		sheet.title = "Relance RH"

		headers = self.create_headers()
		sheet.append(headers)

		for info in information:
			row = [
				info['direction'], info['profile'],
				info['last_name'], info['first_name'],
				info['tel_num'], info['email'],
				info['status'],
			]

			# Interleave dates and managers
			interleaved = []
			for date, manager in zip(info['dates'], info['managers']):
				interleaved.extend([date, manager])
			row.extend(interleaved)
			row.append(info['last_interview'])
			sheet.append(row)

		for column in sheet.columns:
			max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
			adjusted_width = max_length + 3
			sheet.column_dimensions[column[0].column_letter].width = adjusted_width

		sheet.column_dimensions['A'].width = 20

		wb.save(output_path)
		return True

	def extract_path_profile(self, path):
		"""
		This function extracts the path and profile information from the given path.
		path: path to extract information from
		"""
		infos = []

		techno_match = re.search(r' - (.+)$', path)
		if techno_match:
			path_normalized = path.replace('\\', '/')
			profile_info = path_normalized[path_normalized.rfind('/') + 1:]

			infos.append(profile_info)


		path_file_of_user = re.search(r'^([A-Za-z]:/.*?/)', path.replace('\\', '/'))
		if path_file_of_user:
			infos.append(path_file_of_user.group(1))

		return infos

	def verify_value(self, data):
		"""
		This function verifies if the given data (tel number and email ) is valid.
		"""
		tel_num, email = data
		if tel_num and email:
			tel_num = re.sub(r'\D', '', tel_num)
			return self.verify_phone_number(tel_num) and self.verify_email(email)
		return False

	@staticmethod
	def verify_phone_number(phone_number):
		return bool(re.fullmatch(r"0\d{9}", phone_number))

	@staticmethod
	def verify_email(email):
		if email in ExcelOperations.verified_emails:
			return True
		check_email = bool(re.fullmatch(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", email))
		if check_email:
			ExcelOperations.verified_emails.add(email)
			return True
		else:
			if email not in ExcelOperations.errors:
				ExcelOperations.errors.append(email)
			return False

	def verify_format_file_excel(self, file_path):
		wb = load_workbook(file_path)
		sheet = wb.active
		return sheet['C3'].value == "DOSSIER RECRUTEMENT FICHE ENTRETIEN"

	def verify_format_date(self, date):
		if isinstance(date, datetime):
			return date.strftime('%d/%m/%Y')
		if isinstance(date, str) and re.match(r'\d{1,2}/\d{1,2}/\d{2,4}', date):
			return date
		return None

	def format_date(self, date_string):
		if not date_string:
			return None
		try:
			date_object = datetime.strptime(date_string, '%d/%m/%Y')
		except ValueError:
			date_object = datetime.strptime(date_string, '%d/%m/%y')
		return date_object.strftime('%d/%m/%y')

	def find_last_interview(self, dates):
		valid_dates = [datetime.strptime(date, "%d/%m/%y") for date in dates if date]
		return max(valid_dates).strftime("%d/%m/%y") if valid_dates else None

	def is_file_open(self, file_path):
		try:
			os.rename(file_path, file_path)
			return False
		except OSError:
			return True

	def clean_email(self, email):
		if email:
			email = re.sub(r'^(Mail\s*:\s*|:\s*|email\s*)', '', email, flags=re.IGNORECASE).strip()
			email = re.sub(r'\s*([.@])\s*', r'\1', email)
			match = re.search(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', email)
			if match:
				email = match.group(0)
		return email


