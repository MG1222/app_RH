import os
import re
import random
import logging
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook
from datetime import datetime


class ExcelOperations:
	
	def process_excel_files(self, folder_path, progress_bar=None):
		information = []
		excel_files = self.get_excel_files(folder_path)
		total_files = len(excel_files)
		
		for index, (file_path, regex) in enumerate(excel_files):
			try:
				info_from_excel = self.extract_information_from_excel(file_path, regex)
			
				if info_from_excel is None:
					continue
				
				if any(info_from_excel['last_name'] == item['last_name'] and
				       info_from_excel['first_name'] == item['first_name'] for item in information):
					continue
				interviews = []
				for date, manager in zip(info_from_excel['dates'], info_from_excel['managers']):
					interviews.append({'date': date, 'manager': manager})
				info_from_excel['interviews'] = interviews
				information.append(info_from_excel)
			except Exception as e:
				logging.error(f"Error while processing file {file_path}: {e}")
			
			
			if progress_bar:
				progress_value = ((index + 2) / total_files) * 100
				progress_bar['value'] = progress_value
				progress_bar.update()
		
		if not information:
			return None
		
		return information
	
	def get_excel_files(self, folder_path):
		excel_files = []
		for root, dirs, files in os.walk(folder_path):
			for file in files:
				if file.endswith(".xlsx") or file.endswith(".xls"):
					regex = self.extract_path_profile(root)
					file_path = os.path.join(root, file)
					file_size = os.path.getsize(file_path)
					if file_size == 0:
						continue
					check_format = self.verify_format_file_excel(file_path)
					if check_format:
						excel_files.append((file_path, regex))
		return excel_files
	

	
	def extract_contact_information(self, sheet):
		tel_num = sheet['C5'].value
		tel_num = re.sub(r'\D', '', tel_num)
		tel_num_sec = sheet['D5'].value
		email = sheet['E5'].value
		if email is not None:
			email = email.strip().replace("Mail: ", "")
		data_contact = [tel_num, email]
		
		verification_result = self.verify_value(data_contact)
		if not verification_result:
			data_contact = [tel_num_sec, email]
			verification_result = self.verify_value(data_contact)
			if verification_result:
				tel_num = tel_num_sec
		
		tel_num = re.sub(r'\D', '', tel_num)
		tel_num = ' '.join([tel_num[i:i + 2] for i in range(0, len(tel_num), 2)])
		
		return tel_num, email
	
	def extract_interview_information(self, sheet):
		dates = []
		managers = []
		for row in range(7, 10):
			date_value = sheet.cell(row=row, column=8).value
			date_value = self.verify_format_date(date_value)
			if date_value is not None:
				date_value = self.format_date(date_value)
			#	date_value = self.change_date_randomly(date_value)
			dates.append(date_value)
		
		for row in range(7, 10):
			manager_value = sheet.cell(row=row, column=7).value
			managers.append(manager_value if manager_value is not None else "")

			
			last_interview = self.find_last_interview(dates)
			print(last_interview)
			
		
		return dates, managers, last_interview
	
	def extract_information_from_excel(self, file_path, regex):
		wb = load_workbook(file_path)
		sheet = wb.active
		
		tel_num, email = self.extract_contact_information(sheet)
		dates, managers, last_interview = self.extract_interview_information(sheet)
		
		profile = regex[0] if len(regex) > 0 else ''
		direction = file_path
		last_name = sheet['B6'].value
		first_name = sheet['B9'].value
		status = sheet['G22'].value
		if status is None:
			status = sheet['G23'].value
		
		information_perso = {
			'profile': profile,
			'direction': direction,
			'last_name': last_name,
			'first_name': first_name,
			'tel_num': tel_num,
			'email': email,
			'status': status,
			'dates': dates,
			'managers': managers,
			'last_interview': last_interview
		}
		
		return information_perso
	
	def create_headers(self):
		headers = ["repartoire", "profil", "nom", "prenom", "tel", "email", "disponibilite"]
		for i in range(3):
			headers += [f"Date{i + 1}", f"Entretien{i + 1}"]
		headers.append("dernier entretien")
		return [header.upper() for header in headers]
	
	def create_new_excel_file(self, information, output_path):
		if not information:
			raise ValueError("No information to write, file would be empty except headers")
		
		wb = Workbook()
		sheet = wb.active
		sheet.title = "Relance RH"
		
		headers = self.create_headers()
		sheet.append(headers)
		
		for info in information:
			row = [
				info['direction'], info['profile'],
				info['last_name'], info['first_name'],
				info['tel_num'], info['email'],
				info['status'],
			]
			
			for interview in info['interviews']:
				row += [interview['date'], interview['manager']]
			row.append(info['last_interview'])
			sheet.append(row)
			
		
		# column style adjustment
		for column in sheet.columns:
			max_length = 0
			column = [cell for cell in column]
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except:
					pass
			adjusted_width = (max_length + 2)
			sheet.column_dimensions[column[0].column_letter].width = adjusted_width
			repartoire_index = headers.index("REPARTOIRE")
			sheet.column_dimensions[chr(65 + repartoire_index)].width = 15
		wb.save(output_path)
		
		return True
	
	
	def extract_path_profile(self, path):
		infos = []
		techno = re.search(r'- (.+)', path)
		path_file_of_user = re.search(r'(.+)\\', path)
		if techno:
			infos.append(techno.group(1))
		if path_file_of_user:
			infos.append(path_file_of_user.group(1))
		
		return infos
	
	
	def verify_value(self, data):
		tel_num, email = data
		if tel_num is not None and email is not None:
			tel_num = re.sub(r'\D', '', tel_num)
			return self.verify_phone_number(tel_num) and self.verify_email(email)
		else:
			return False
	
	
	@staticmethod
	def verify_phone_number(phone_number):
		return bool(re.fullmatch(r"0\d{9}", phone_number))
	
	
	@staticmethod
	def verify_email(email):
		return bool(re.fullmatch(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", email))
	
	
	def verify_format_file_excel(self, file_path):
		wb = load_workbook(file_path)
		sheet = wb.active
		title = "DOSSIER RECRUTEMENT FICHE ENTRETIEN"
		if sheet['C3'].value != title:
			return False
		return True
	
	
	def verify_format_date(self, date):
		if isinstance(date, datetime):
			return date.strftime('%d/%m/%Y')
		if isinstance(date, str) and re.match(r'\d{1,2}/\d{1,2}/\d{2,4}', date):
			return date
		if date == '___/___/__':
			return None
		return None
	
	
	def format_date(self, date_string):
		# 'dd/mm/yyyy' -> 'dd/mm/yy'
		try:
			date_object = datetime.strptime(date_string, '%d/%m/%Y')
		except ValueError:
			# 'dd/mm/yy' -> 'dd/mm/yy'
			date_object = datetime.strptime(date_string, '%d/%m/%y')
		
		# 'dd/mm/yy' -> 'dd/mm/yy'
		formatted_date = date_object.strftime('%d/%m/%y')
		
		return formatted_date
	
	def find_last_interview(self, dates):
		for date in dates:
			if date is not None:
				valid_dates = [datetime.strptime(date, "%d/%m/%y") for date in dates if date is not None]
				if valid_dates:
					most_recent_date = max(valid_dates)
					most_recent_date = most_recent_date.strftime("%d/%m/%y")
					return most_recent_date
				else:
					return None
